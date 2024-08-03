import os
import re
from openpyxl import load_workbook

from Managers.filemanager import FileManager

class ExcelFileManager(FileManager):
    def __init__(self, path_file: str) -> None:
        super().__init__(path_file)
        self.workbook = load_workbook(self.path_file)
        self.extension = 'xlsx'

    def split_file(self) -> None:
        output_dir = f"{self.file_name}"
        os.makedirs(output_dir, exist_ok=True)
        
        for sheet in self.workbook.sheetnames:
            new_wb = load_workbook(self.path_file)
            new_ws = new_wb[sheet]
            
            output_xlsx = f"{output_dir}/{sheet}.{self.extension}"
            new_wb.save(output_xlsx)

    def merge_files(self, output_xlsx: str) -> None:
        merged_wb = load_workbook(self.path_file)
        input_dir = f"{self.file_name}"
        page_files = [f"{input_dir}/{file}" for file in os.listdir(input_dir) if file.endswith('.xlsx')]
        page_files.sort(key=lambda f: int(re.search(r'(\d+).xlsx', f).group(1)))

        for file in page_files:
            wb = load_workbook(file)
            for sheet in wb.sheetnames:
                ws = wb[sheet]
                merged_wb.create_sheet(sheet)
                merged_ws = merged_wb[sheet]
                
                for row in ws.iter_rows():
                    for cell in row:
                        merged_ws[cell.coordinate].value = cell.value
        
        merged_wb.save(f'{output_xlsx}_merged.{self.extension}')
        print(f"Excel files merged into {output_xlsx}_merged.{self.extension}'")
